#!/usr/bin/env python3
from __future__ import annotations

import argparse, csv, itertools, json, os, re, sqlite3, urllib.request, zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = "https://www.bungie.net"
VALID_TIERS = {"S", "A", "B", "C", "D", "E"}

# These are not real exact perk names in the manifest when used by the sheet as a generic mag family.
# Treat them as "no specific mag requirement" for Barrel-Mag files rather than causing the row to fail.
GENERIC_MAG_TERMS = {
    "battery",
}

# Suffixes the sheet uses for context, but Bungie's manifest usually doesn't include in the display name.
WEAPON_SUFFIX_PATTERNS = [
    r"\s+BRAVE\s+version$",
    r"\s+RotN\s+version$",
    r"\s+Adept\s+version$",
    r"\s+Harrowed\s+version$",
    r"\s+Timelost\s+version$",
    r"\s+Pantheon\s+version$",
]


def normalize_quotes(x):
    if x is None:
        return ""
    return str(x).strip().replace("\u2019", "'").replace("\u2018", "'").replace("\u201c", '"').replace("\u201d", '"')


def clean(x):
    s = normalize_quotes(x)
    return re.sub(r"\s+", " ", s).strip()


def norm(x):
    return clean(x).casefold()


def weapon_name_candidates(name):
    """
    Try exact sheet name first, then strip contextual suffixes like:
      "Succession BRAVE version" -> "Succession"
      "Cold Comfort RotN version" -> "Cold Comfort"
    """
    base = clean(name)
    candidates = [base]

    for pattern in WEAPON_SUFFIX_PATTERNS:
        stripped = re.sub(pattern, "", base, flags=re.I).strip()
        if stripped and stripped not in candidates:
            candidates.append(stripped)

    # Also handle parenthetical variants just in case.
    stripped_paren = re.sub(r"\s*\([^)]*\)\s*$", "", base).strip()
    if stripped_paren and stripped_paren not in candidates:
        candidates.append(stripped_paren)

    return candidates


def split_perks(x):
    raw = normalize_quotes(x)
    if not raw:
        return []
    raw = re.sub(r"\bEnhanced\b", "", raw, flags=re.I)
    out, seen = [], set()
    for p in re.split(r"[\n\r,;]+", raw):
        p = clean(p)
        if not p or norm(p) in {"none", "n/a", "na", "-", "?", "need testing"}:
            continue
        k = norm(p)
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out


def split_mag(x):
    # Preserve specific mags, but drop generic category words like "Battery".
    mags = []
    for p in split_perks(x):
        if norm(p) in GENERIC_MAG_TERMS:
            continue
        mags.append(p)
    return mags


def find_header(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return None

    for r in range(1, min(max_row, 30) + 1):
        h = {}
        for c in range(1, max_col + 1):
            v = clean(ws.cell(r, c).value)
            if v:
                h[norm(v)] = c
        if all(x in h for x in ["name", "perk 1", "perk 2", "tier"]):
            return r, h
    return None


def col(h, *names):
    for n in names:
        if norm(n) in h:
            return h[norm(n)]
    return None


def read_rows(xlsx):
    """
    Read Aegis rows from sheets.

    Handles two common spreadsheet shapes:
    1) Normal rows where Name, Tier, Rank, Perk 1, Perk 2 are all on the same row.
    2) Continuation/merged-style rows where Tier/Rank appear on the first visual row,
       but the actual alternate-version perk row below has blank Tier/Rank cells.
       Example:
         Zaouli's Bane                 Tier A / Rank 6 / no perks
         Zaouli's Bane Pantheon version blank tier/rank / Firefly + Chaos Reshaped
    """
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
    rows = []

    for ws in wb.worksheets:
        found = find_header(ws)
        if not found:
            continue

        hr, h = found
        c_name, c_tier = col(h, "Name"), col(h, "Tier")
        c_rank, c_p1, c_p2 = col(h, "Rank"), col(h, "Perk 1"), col(h, "Perk 2")
        c_barrel, c_mag = col(h, "Barrel"), col(h, "Mag")
        c_origin, c_notes = col(h, "Origin Trait", "Origin Trai"), col(h, "Notes")

        max_row = ws.max_row or 0
        last_rank = ""
        last_tier = ""

        for r in range(hr + 1, max_row + 1):
            name = clean(ws.cell(r, c_name).value)
            raw_rank = clean(ws.cell(r, c_rank).value) if c_rank else ""
            raw_tier = clean(ws.cell(r, c_tier).value).upper()

            # Carry down rank/tier for visual continuation rows.
            # This is needed because some Aegis sheet sections put alternate versions
            # on the line below the ranked base row.
            rank = raw_rank or last_rank
            tier = raw_tier or last_tier

            if raw_rank:
                last_rank = raw_rank
            if raw_tier in VALID_TIERS:
                last_tier = raw_tier

            if not name or tier not in VALID_TIERS:
                continue

            p1 = split_perks(ws.cell(r, c_p1).value)
            p2 = split_perks(ws.cell(r, c_p2).value)
            if not p1 or not p2:
                # Still carry tier/rank, but do not emit a wishlist row without main perks.
                continue

            rows.append({
                "sheet": ws.title,
                "row": r,
                "name": name,
                "tier": tier,
                "rank": rank,
                "p1": p1,
                "p2": p2,
                "barrel": split_perks(ws.cell(r, c_barrel).value) if c_barrel else [],
                "mag": split_mag(ws.cell(r, c_mag).value) if c_mag else [],
                "origin": clean(ws.cell(r, c_origin).value) if c_origin else "",
                "notes": clean(ws.cell(r, c_notes).value) if c_notes else "",
            })

    return rows

def get_json(url, key=None):
    headers = {"User-Agent": "aegis-dim-builder/1.4"}
    if key:
        headers["X-API-Key"] = key
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req) as res:
        return json.loads(res.read().decode("utf-8"))


def download(url, path, key=None):
    headers = {"User-Agent": "aegis-dim-builder/1.4"}
    if key:
        headers["X-API-Key"] = key
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req) as res:
        path.write_bytes(res.read())


def manifest_db(cache):
    key = os.environ.get("BUNGIE_API_KEY")
    cache.mkdir(parents=True, exist_ok=True)
    data = get_json(ROOT + "/Platform/Destiny2/Manifest/", key)
    resp = data.get("Response", data)
    rel = resp["mobileWorldContentPaths"]["en"]
    zpath = cache / "manifest.zip"
    download(ROOT + rel, zpath, key)
    with zipfile.ZipFile(zpath) as z:
        name = next((n for n in z.namelist() if n.endswith((".content", ".sqlite", ".db"))), z.namelist()[0])
        z.extract(name, cache)
    return cache / name


def iter_items(db):
    con = sqlite3.connect(db)
    try:
        tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        table = "DestinyInventoryItemDefinition" if "DestinyInventoryItemDefinition" in tables else next(t for t in tables if "InventoryItemDefinition" in t)
        for (txt,) in con.execute(f"SELECT json FROM {table}"):
            try:
                yield json.loads(txt)
            except Exception:
                pass
    finally:
        con.close()


def build_index(db):
    weapons, plugs = defaultdict(list), defaultdict(list)
    for it in iter_items(db):
        name = clean((it.get("displayProperties") or {}).get("name"))
        if not name or "hash" not in it:
            continue
        h = int(it["hash"])
        if int(it.get("itemType", -1)) == 3:
            weapons[norm(name)].append(h)
        if it.get("plug") is not None:
            plugs[norm(name)].append(h)
    return {k: sorted(set(v)) for k, v in weapons.items()}, {k: sorted(set(v)) for k, v in plugs.items()}


def find_weapon_hashes(row, weapons, missing):
    for candidate in weapon_name_candidates(row["name"]):
        wh = weapons.get(norm(candidate), [])
        if wh:
            return wh

    missing.append([row["sheet"], row["row"], row["name"], "weapon", " / ".join(weapon_name_candidates(row["name"])), row["tier"], row["rank"]])
    return []


def note(row):
    bits = [
        f"{row['name']} - Rank {row['rank']}, Tier {row['tier']} by TheAegisRelic",
        f"Sheet: {row['sheet']}",
        "Column 1: " + ", ".join(row["p1"]),
        "Column 2: " + ", ".join(row["p2"]),
    ]
    if row["barrel"]:
        bits.append("Barrel: " + ", ".join(row["barrel"]))
    if row["mag"]:
        bits.append("Mag: " + ", ".join(row["mag"]))
    else:
        bits.append("Mag: Any / generic")
    if row["origin"]:
        bits.append("Origin Trait: " + row["origin"])
    if row["notes"]:
        bits.append("Notes: " + row["notes"])
    return "\\n".join(bits).replace("#", "").replace("\r", "")


def hashes_for_names(names, plugs, row, missing, missing_type):
    all_hashes = []
    for name in names:
        hs = plugs.get(norm(name), [])
        if hs:
            all_hashes.extend(hs[:20])
        else:
            missing.append([row["sheet"], row["row"], row["name"], missing_type, name, row["tier"], row["rank"]])
    return sorted(set(all_hashes))


def generate(rows, weapons, plugs, tiers, require_barrel_mag=False):
    lines, missing, seen = [], [], set()
    for row in rows:
        if row["tier"] not in tiers:
            continue

        wh = find_weapon_hashes(row, weapons, missing)
        if not wh:
            continue

        p1s = hashes_for_names(row["p1"], plugs, row, missing, "perk_1")
        p2s = hashes_for_names(row["p2"], plugs, row, missing, "perk_2")
        if not p1s or not p2s:
            continue

        if require_barrel_mag:
            barrels = hashes_for_names(row["barrel"], plugs, row, missing, "barrel") if row["barrel"] else []
            mags = hashes_for_names(row["mag"], plugs, row, missing, "mag") if row["mag"] else []

            # If the sheet has specific barrel/mag names, require them.
            # If the sheet only has a generic mag family like Battery, skip that requirement.
            # This avoids dropping fusion/trace/LFR rows from Barrel-Mag outputs.
            combo_parts = []
            if barrels:
                combo_parts.append(barrels)
            if mags:
                combo_parts.append(mags)
            combo_parts.extend([p1s, p2s])
            combo_list = list(itertools.product(*combo_parts))
        else:
            combo_list = list(itertools.product(p1s, p2s))

        lines += ["", "//notes:" + note(row)]

        for w in wh[:20]:
            for combo in combo_list:
                perk_hashes = ",".join(str(x) for x in combo)
                line = f"dimwishlist:item={w}&perks={perk_hashes}"
                if line not in seen:
                    seen.add(line)
                    lines.append(line)

    return lines, missing


def write_file(path, title, desc, lines, require_barrel_mag=False):
    strict = "weapon + one recommended Perk 1 + one recommended Perk 2"
    if require_barrel_mag:
        strict = "weapon + recommended barrel/mag when specific + one recommended Perk 1 + one recommended Perk 2"

    head = [
        "title:" + title,
        "description:" + desc,
        "// Generated from current Aegis Endgame Analysis Excel.",
        f"// Strict mode: {strict}.",
        "",
    ]
    Path(path).write_text("\n".join(head + lines).strip() + "\n", encoding="utf-8")


def build_outputs(rows, weapons, plugs, outdir, require_barrel_mag, suffix, desc_extra):
    all_missing = []
    outputs = [
        ("Aegis-Endgame-S", {"S"}, "Aegis Endgame S-tier", "Current Aegis Endgame S-tier weapons."),
        ("Aegis-Endgame-A", {"A"}, "Aegis Endgame A-tier", "Current Aegis Endgame A-tier weapons."),
        ("Aegis-Endgame-A-and-S", {"A", "S"}, "Aegis Endgame A and S-tier", "Current Aegis Endgame A/S-tier weapons."),
        ("Aegis-Endgame-A-through-E", {"A", "B", "C", "D", "E"}, "Aegis Endgame A through E-tier", "Current Aegis Endgame A-E-tier weapons."),
        ("Aegis-Endgame-C-through-E", {"C", "D", "E"}, "Aegis Endgame C through E-tier", "Current Aegis Endgame C-E-tier weapons."),
        ("Aegis-Endgame-S-through-E", {"S", "A", "B", "C", "D", "E"}, "Aegis Endgame S through E-tier", "Current Aegis Endgame S-E-tier weapons."),
    ]

    for base_name, tiers, title, desc in outputs:
        fn = f"{base_name}{suffix}.txt"
        title = f"{title}{desc_extra}"
        desc = f"{desc.rstrip('.')}{desc_extra}."
        lines, missing = generate(rows, weapons, plugs, tiers, require_barrel_mag=require_barrel_mag)
        write_file(Path(outdir) / fn, title, desc, lines, require_barrel_mag=require_barrel_mag)
        all_missing += missing
        print(f"Wrote {fn}: {sum(1 for l in lines if l.startswith('dimwishlist:'))} DIM lines; {len(missing)} unresolved")

    return all_missing


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="data.xlsx")
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--cache-dir", default=".cache")
    args = ap.parse_args()

    rows = read_rows(Path(args.excel))
    print(f"Found {len(rows)} S/A/B/C/D/E rows")
    if not rows:
        raise SystemExit("No ranked rows found. Check that the Excel workbook has Name, Perk 1, Perk 2, Tier headers.")

    db = manifest_db(Path(args.cache_dir))
    weapons, plugs = build_index(db)
    print(f"Indexed {len(weapons)} weapon names and {len(plugs)} plug names")

    all_missing = []
    all_missing += build_outputs(rows, weapons, plugs, args.outdir, False, "", "")
    all_missing += build_outputs(rows, weapons, plugs, args.outdir, True, "-Barrel-Mag", " with Barrel and Mag")

    # Deduplicate unresolved rows for readability.
    deduped = []
    seen = set()
    for row in all_missing:
        key = tuple(row)
        if key not in seen:
            seen.add(key)
            deduped.append(row)

    with open(Path(args.outdir) / "Aegis-Unresolved-Names.csv", "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f)
        wr.writerow(["sheet", "row", "weapon", "type", "missing", "tier", "rank"])
        wr.writerows(deduped)


if __name__ == "__main__":
    main()
